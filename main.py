#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
加密Excel合并工具 v2.0
- 支持拖拽/选择多个加密Excel文件
- 密码按城市区分，自动匹配
- 支持大文件(~1GB)处理，内存流式优化
- 跨平台支持 macOS / Windows / Linux
- 新增：并行解密、数据去重、预览、日志导出、命令行模式
"""

import sys
import os
import io
import struct
import hashlib
import hmac
import tempfile
import time
import json
import csv
import logging
from collections import OrderedDict
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import List, Tuple, Dict, Optional, Any
from dataclasses import dataclass
import platform
import shutil


def _get_system_font():
    """获取跨平台中文字体"""
    system = platform.system()
    if system == 'Darwin':
        return 'PingFang SC'
    elif system == 'Windows':
        return 'Microsoft YaHei'
    else:
        return 'Noto Sans CJK SC'


def _get_mono_font():
    """获取跨平台等宽字体"""
    system = platform.system()
    if system == 'Darwin':
        return 'Menlo'
    elif system == 'Windows':
        return 'Consolas'
    else:
        return 'DejaVu Sans Mono'



from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QPushButton, QTableWidget, QTableWidgetItem, QFileDialog,
    QProgressBar, QTextEdit, QGroupBox, QHeaderView, QComboBox,
    QStyledItemDelegate, QAbstractItemView, QMessageBox, QLineEdit,
    QSpinBox, QCheckBox
)
from PyQt6.QtCore import Qt, pyqtSignal, QThread
from PyQt6.QtGui import QColor, QFont, QDragEnterEvent, QDropEvent

import olefile
from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes
from cryptography.hazmat.backends import default_backend
import openpyxl

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

# 尝试导入 msoffcrypto 库（PyPI 包名已从 msoffcrypto-python 更名为 msoffcrypto）
_HAS_MSOFFCRYPTO = False
try:
    import msoffcrypto
    _HAS_MSOFFCRYPTO = True
except ImportError:
    pass


# ============================================================
# Excel 解密模块
# ============================================================

@dataclass
class FileInfo:
    """文件信息数据类"""
    filepath: str
    city: str = ""
    password: str = ""
    target_sheet: str = ""
    
class DecryptionError(Exception):
    """解密异常"""
    pass

class MergeError(Exception):
    """合并异常"""
    pass


def _hash_func(hash_algorithm):
    """根据算法名返回 hashlib 哈希函数"""
    algo_map = {
        'SHA1': hashlib.sha1,
        'SHA256': hashlib.sha256,
        'SHA384': hashlib.sha384,
        'SHA512': hashlib.sha512,
        'MD5': hashlib.md5,
    }
    name = hash_algorithm.upper().replace('-', '').replace('_', '')
    if name not in algo_map:
        raise DecryptionError(f"不支持的哈希算法: {hash_algorithm}")
    return algo_map[name]


def _b64decode(value):
    """安全解码 base64 字符串（XML 中的二进制值是 base64 编码，不是 hex）"""
    import base64
    if not value:
        return b''
    value = value.strip().replace('\n', '').replace('\r', '').replace(' ', '')
    padding = 4 - len(value) % 4
    if padding != 4:
        value += '=' * padding
    return base64.b64decode(value)


def _aes_cbc_decrypt(key, iv, data):
    """AES-CBC 解密"""
    cipher = Cipher(algorithms.AES(key), modes.CBC(iv), backend=default_backend())
    decryptor = cipher.decryptor()
    decrypted = decryptor.update(data) + decryptor.finalize()
    return decrypted


def _remove_pkcs5_padding(data):
    """移除 PKCS5/7 填充"""
    if not data:
        return data
    pad_len = data[-1]
    if pad_len < 1 or pad_len > 16:
        return data
    # 验证填充
    if data[-pad_len:] != bytes([pad_len]) * pad_len:
        return data
    return data[:-pad_len]


def decrypt_file(filepath: str, password: str) -> bytes:
    """
    解密加密的 Excel 文件。
    优先使用 msoffcrypto 库，不可用时回退到内嵌实现。
    
    Args:
        filepath: 文件路径
        password: 解密密码
        
    Returns:
        解密后的字节数据
        
    Raises:
        DecryptionError: 解密失败
    """
    if not password:
        raise DecryptionError("密码不能为空")

    # 优先使用 msoffcrypto 库
    if _HAS_MSOFFCRYPTO:
        return _decrypt_with_msoffcrypto(filepath, password)

    # 回退到内嵌实现
    return _decrypt_embedded(filepath, password)


def _decrypt_with_msoffcrypto(filepath, password):
    """使用 msoffcrypto 库解密"""
    try:
        with open(filepath, 'rb') as f:
            file = msoffcrypto.OfficeFile(f)
            buf = io.BytesIO()
            file.load_key(password=password)
            file.decrypt(buf)
            return buf.getvalue()
    except msoffcrypto.InvalidKeyError:
        raise DecryptionError("密码错误")
    except Exception as e:
        err_msg = str(e).lower()
        if 'not encrypted' in err_msg or 'not an ole' in err_msg:
            with open(filepath, 'rb') as f:
                return f.read()
        raise DecryptionError(f"解密失败: {str(e)}")


def _decrypt_embedded(filepath, password):
    """内嵌解密实现"""
    if not olefile.isOleFile(filepath):
        with open(filepath, 'rb') as f:
            return f.read()

    ole = olefile.OleFileIO(filepath)

    if not ole.exists('EncryptionInfo') or not ole.exists('EncryptedPackage'):
        ole.close()
        with open(filepath, 'rb') as f:
            return f.read()

    encryption_info = ole.openstream('EncryptionInfo').read()
    encrypted_package = ole.openstream('EncryptedPackage').read()
    ole.close()

    version_major = struct.unpack('<H', encryption_info[0:2])[0]
    version_minor = struct.unpack('<H', encryption_info[2:4])[0]

    if version_major == 4:
        return _decrypt_agile(encryption_info, encrypted_package, password)
    elif version_major in (2, 3):
        return _decrypt_standard(encryption_info, encrypted_package, password)
    else:
        raise DecryptionError(f"不支持的加密版本: {version_major}.{version_minor}")


def _decrypt_agile(encryption_info, encrypted_package, password):
    """Agile Encryption 解密 (Office 2010+) - 修正版"""
    import xml.etree.ElementTree as ET

    # EncryptionInfo 头部: 2B major + 2B minor + 4B reserved = 8 bytes
    xml_data = encryption_info[8:]
    xml_str = xml_data.decode('utf-8', errors='ignore').rstrip('\x00')

    if not xml_str.startswith('<'):
        idx = xml_str.find('<')
        if idx > 0:
            xml_str = xml_str[idx:]

    try:
        root = ET.fromstring(xml_str)
    except ET.ParseError as e:
        raise DecryptionError(f"解析加密信息XML失败: {e}")

    # 提取命名空间
    ns = ''
    if root.tag.startswith('{'):
        ns = root.tag.split('}')[0] + '}'

    # ---- 解析 keyData ----
    key_data = root.find(f'{ns}keyData')
    if key_data is None:
        raise DecryptionError("未找到 keyData 元素")

    # IMPORTANT: XML 中的二进制值是 base64 编码，不是 hex
    kd_salt = _b64decode(key_data.get('saltValue', ''))
    kd_hash_algo = key_data.get('hashAlgorithm', 'SHA512')
    kd_key_bits = int(key_data.get('keyBits', '256'))
    kd_cipher_chaining = key_data.get('cipherChaining', 'ChainingModeCBC')

    # ---- 解析 encryptedKey (password encryptor) ----
    key_encryptors = root.find(f'{ns}keyEncryptors')
    if key_encryptors is None:
        raise DecryptionError("未找到 keyEncryptors 元素")

    encrypted_key = None
    for ke in key_encryptors:
        ek = ke.find(f'{ns}encryptedKey')
        if ek is not None:
            encrypted_key = ek
            break
        if 'spinCount' in ke.attrib:
            encrypted_key = ke
            break

    if encrypted_key is None:
        raise DecryptionError("未找到 encryptedKey 元素")

    ek_spin_count = int(encrypted_key.get('spinCount', '100000'))
    ek_salt = _b64decode(encrypted_key.get('saltValue', ''))
    ek_hash_algo = encrypted_key.get('hashAlgorithm', 'SHA512')
    ek_key_bits = int(encrypted_key.get('keyBits', '256'))
    ek_encrypted_key_value = _b64decode(encrypted_key.get('encryptedKeyValue', ''))
    ek_encrypted_verifier_hash_input = _b64decode(encrypted_key.get('encryptedVerifierHashInput', ''))
    ek_encrypted_verifier_hash_value = _b64decode(encrypted_key.get('encryptedVerifierHashValue', ''))

    # ---- 密钥推导 (MS-OFFCRYPTO 2.3.4.7) ----
    hash_func = _hash_func(ek_hash_algo)
    password_bytes = password.encode('UTF-16LE')
    h = hash_func(ek_salt + password_bytes).digest()
    for _ in range(ek_spin_count):
        h = hash_func(h).digest()
    derived_key = h[:ek_key_bits // 8]

    # ---- 使用 HMAC 推导各用途的加密密钥 ----
    def _derive_encryption_key(base_key, block_index):
        """使用 HMAC 推导指定用途的加密密钥"""
        block_bytes = struct.pack('<I', block_index)
        derived = hmac.new(base_key, block_bytes, hash_func).digest()
        return derived[:ek_key_bits // 8]

    # ---- 验证密码 ----
    verifier_key = _derive_encryption_key(derived_key, 0x00000000)

    def _make_iv(salt):
        if len(salt) >= 16:
            return salt[:16]
        return salt + b'\x00' * (16 - len(salt))

    iv = _make_iv(ek_salt)

    try:
        verifier_input = _aes_cbc_decrypt(verifier_key, iv, ek_encrypted_verifier_hash_input)
        verifier_hash = _aes_cbc_decrypt(verifier_key, iv, ek_encrypted_verifier_hash_value)
    except Exception as e:
        raise DecryptionError(f"密码验证解密失败: {e}")

    computed_hash = hash_func(_remove_pkcs5_padding(verifier_input)).digest()
    expected_hash = _remove_pkcs5_padding(verifier_hash)
    if computed_hash[:len(expected_hash)] != expected_hash[:len(computed_hash)]:
        raise DecryptionError("密码错误")

    # ---- 解密 secret key ----
    key_encrypt_key = _derive_encryption_key(derived_key, 0x00000001)
    try:
        secret_key = _aes_cbc_decrypt(key_encrypt_key, iv, ek_encrypted_key_value)
        secret_key = _remove_pkcs5_padding(secret_key)
    except Exception as e:
        raise DecryptionError(f"解密密钥失败: {e}")

    # ---- 解密 EncryptedPackage ----
    # 前 4 字节是原始数据大小 (little-endian uint32)
    if len(encrypted_package) < 4:
        raise DecryptionError("加密数据包太小")

    original_size = struct.unpack('<I', encrypted_package[:4])[0]
    encrypted_data = encrypted_package[4:]

    # 确保数据长度是 16 的倍数
    if len(encrypted_data) % 16 != 0:
        encrypted_data = encrypted_data[:len(encrypted_data) - (len(encrypted_data) % 16)]

    # Package 密钥推导
    package_key = _derive_encryption_key(secret_key, 0x00000005)
    package_iv = _make_iv(kd_salt)

    try:
        decrypted = _aes_cbc_decrypt(package_key, package_iv, encrypted_data)
    except Exception as e:
        raise DecryptionError(f"解密数据包失败: {e}")

    decrypted = _remove_pkcs5_padding(decrypted)

    if original_size > 0 and original_size < len(decrypted):
        decrypted = decrypted[:original_size]

    # 验证结果是否为有效 ZIP 文件
    if decrypted[:2] != b'PK':
        # 尝试不带 HMAC key derivation 的简单方式
        try:
            return _decrypt_agile_simple(encryption_info, encrypted_package, password)
        except Exception:
            raise DecryptionError("解密结果不是有效的Excel文件")

    return decrypted


def _decrypt_agile_simple(encryption_info, encrypted_package, password):
    """Agile Encryption 简化解密（不使用 HMAC key derivation，兼容某些实现）"""
    import xml.etree.ElementTree as ET

    xml_data = encryption_info[8:]
    xml_str = xml_data.decode('utf-8', errors='ignore').rstrip('\x00')
    if not xml_str.startswith('<'):
        idx = xml_str.find('<')
        if idx > 0:
            xml_str = xml_str[idx:]
    root = ET.fromstring(xml_str)

    ns = ''
    if root.tag.startswith('{'):
        ns = root.tag.split('}')[0] + '}'

    key_data = root.find(f'{ns}keyData')
    kd_salt = _b64decode(key_data.get('saltValue', ''))

    key_encryptors = root.find(f'{ns}keyEncryptors')
    encrypted_key = None
    for ke in key_encryptors:
        ek = ke.find(f'{ns}encryptedKey')
        if ek is not None:
            encrypted_key = ek
            break
        if 'spinCount' in ke.attrib:
            encrypted_key = ke
            break

    ek_spin_count = int(encrypted_key.get('spinCount', '100000'))
    ek_salt = _b64decode(encrypted_key.get('saltValue', ''))
    ek_hash_algo = encrypted_key.get('hashAlgorithm', 'SHA512')
    ek_key_bits = int(encrypted_key.get('keyBits', '256'))
    ek_encrypted_key_value = _b64decode(encrypted_key.get('encryptedKeyValue', ''))
    ek_encrypted_verifier_hash_input = _b64decode(encrypted_key.get('encryptedVerifierHashInput', ''))
    ek_encrypted_verifier_hash_value = _b64decode(encrypted_key.get('encryptedVerifierHashValue', ''))

    # 简单密钥推导（直接截取，不用 HMAC）
    hash_func = _hash_func(ek_hash_algo)
    h = hash_func(ek_salt + password.encode('UTF-16LE')).digest()
    for _ in range(ek_spin_count):
        h = hash_func(h).digest()
    derived_key = h[:ek_key_bits // 8]

    # 验证
    iv = (ek_salt + b'\x00' * 16)[:16]
    verifier_input = _aes_cbc_decrypt(derived_key, iv, ek_encrypted_verifier_hash_input)
    verifier_hash = _aes_cbc_decrypt(derived_key, iv, ek_encrypted_verifier_hash_value)

    computed = hash_func(_remove_pkcs5_padding(verifier_input)).digest()
    expected = _remove_pkcs5_padding(verifier_hash)
    if computed[:len(expected)] != expected[:len(computed)]:
        raise DecryptionError("密码错误")

    # 解密 secret key
    secret_key = _remove_pkcs5_padding(_aes_cbc_decrypt(derived_key, iv, ek_encrypted_key_value))

    # 解密 package
    original_size = struct.unpack('<I', encrypted_package[:4])[0]
    encrypted_data = encrypted_package[4:]
    if len(encrypted_data) % 16 != 0:
        encrypted_data = encrypted_data[:len(encrypted_data) - (len(encrypted_data) % 16)]

    package_iv = (kd_salt + b'\x00' * 16)[:16]
    decrypted = _remove_pkcs5_padding(_aes_cbc_decrypt(secret_key, package_iv, encrypted_data))

    if original_size > 0 and original_size < len(decrypted):
        decrypted = decrypted[:original_size]

    return decrypted


def _decrypt_standard(encryption_info, encrypted_package, password):
    """Standard ECMA-376 Encryption 解密 (Office 2007)"""
    # EncryptionInfo 格式:
    # [0-1] Major version (2 or 3)
    # [2-3] Minor version
    # [4-7] Reserved
    # [8-11] Encryption header size
    # [12-15] Encryption header flags
    # [16-19] Size extra
    # [20-23] Algorithm ID (0x00006610 = AES-128, 0x00006611 = AES-192, 0x00006612 = AES-256)
    # [24-27] Hash ID (0x00008004 = SHA1)
    # [28-31] Key size
    # [32-35] Provider type
    # [36-39] Reserved 1
    # [40-43] Reserved 2
    # [44-..] CSP name (UTF-16LE, null-terminated)

    if len(encryption_info) < 44:
        raise DecryptionError("加密信息太短")

    flags = struct.unpack('<I', encryption_info[12:16])[0]
    algo_id = struct.unpack('<I', encryption_info[20:24])[0]
    hash_id = struct.unpack('<I', encryption_info[24:28])[0]
    key_size = struct.unpack('<I', encryption_info[28:32])[0]

    # 算法映射
    algo_map = {
        0x660E: ('AES', 128),
        0x660F: ('AES', 192),
        0x6610: ('AES', 256),
    }
    hash_map = {
        0x8003: 'SHA1',
        0x8004: 'SHA256',
        0x800C: 'SHA512',
    }

    # 处理 algo_id
    if algo_id in algo_map:
        _, default_bits = algo_map[algo_id]
        if key_size == 0:
            key_size = default_bits
    else:
        key_size = key_size or 128

    hash_algo = hash_map.get(hash_id, 'SHA1')

    # 解析加密数据 (在 EncryptionInfo 头部之后)
    # Salt + verifier + verifierHash
    header_size = struct.unpack('<I', encryption_info[8:12])[0]
    enc_data_start = 44 + 2  # 跳过 CSP name 的最小长度

    # 找到 salt 的位置 - 在 header 后面
    # 格式: salt(16) + encryptedVerifier(16) + encryptedVerifierHash(32)
    salt_size = 16
    verifier_size = 16
    verifier_hash_size = 32  # SHA1 = 20 bytes, 但加密后为 32 bytes (AES block)

    # 读取 salt + verifier + verifier hash
    offset = 44  # 跳过固定头部
    # CSP 名称是 UTF-16LE，找到结尾
    while offset < len(encryption_info) - 1:
        if encryption_info[offset:offset + 2] == b'\x00\x00':
            offset += 2
            break
        offset += 2

    # 现在 offset 指向 salt
    if offset + salt_size + verifier_size + verifier_hash_size > len(encryption_info):
        # 回退到标准偏移
        offset = 44 + 2  # 保守偏移

    salt = encryption_info[offset:offset + salt_size]
    offset += salt_size
    verifier = encryption_info[offset:offset + verifier_size]
    offset += verifier_size
    verifier_hash = encryption_info[offset:offset + verifier_hash_size]

    # 推导密钥
    password_bytes = password.encode('UTF-16LE')
    hash_func = _hash_func(hash_algo)
    h = hash_func(salt + password_bytes).digest()
    # 使用 deriveKey 步骤
    # MS-OFFCRYPTO: deriveKey
    derived = h[:key_size // 8]
    if len(derived) < key_size // 8:
        # 如果哈希输出不够长，需要扩展
        derived = h + b'\x00' * (key_size // 8 - len(derived))
    derived = derived[:key_size // 8]

    # 验证密码
    try:
        decrypted_verifier = _aes_cbc_decrypt(derived, salt, verifier)
    except Exception:
        raise DecryptionError("密码错误或解密失败")

    try:
        decrypted_verifier_hash = _aes_cbc_decrypt(derived, salt, verifier_hash)
    except Exception:
        raise DecryptionError("密码错误或解密失败")

    # 验证
    computed_hash = hash_func(decrypted_verifier).digest()
    expected = _remove_pkcs5_padding(decrypted_verifier_hash)
    if computed_hash[:len(expected)] != expected[:len(computed_hash)]:
        raise DecryptionError("密码错误")

    # 解密数据包
    # encryptedPackage 前4字节是未加密数据的大小
    if len(encrypted_package) < 4:
        raise DecryptionError("加密数据包太小")

    original_size = struct.unpack('<I', encrypted_package[:4])[0]
    encrypted_data = encrypted_package[4:]

    # 确保数据长度是16的倍数
    if len(encrypted_data) % 16 != 0:
        encrypted_data = encrypted_data[:len(encrypted_data) - (len(encrypted_data) % 16)]

    try:
        decrypted = _aes_cbc_decrypt(derived, salt, encrypted_data)
    except Exception:
        raise DecryptionError("解密数据失败")

    decrypted = _remove_pkcs5_padding(decrypted)

    # 截取到原始大小
    if original_size > 0 and original_size < len(decrypted):
        decrypted = decrypted[:original_size]

    return decrypted


def is_encrypted_file(filepath: str) -> bool:
    """
    检查文件是否为加密的 Excel 文件
    
    Args:
        filepath: 文件路径
        
    Returns:
        是否加密
    """
    try:
        if not olefile.isOleFile(filepath):
            return False
        ole = olefile.OleFileIO(filepath)
        result = ole.exists('EncryptionInfo')
        ole.close()
        return result
    except Exception:
        return False


def decrypt_to_tempfile(filepath: str, password: str) -> str:
    """
    解密 Excel 文件到临时文件。
    优先使用流式写入（msoffcrypto 文件到文件），节省大文件内存。
    
    Args:
        filepath: 源文件路径
        password: 解密密码
        
    Returns:
        临时文件路径
        
    Raises:
        DecryptionError: 解密失败
    """
    fd, temp_path = tempfile.mkstemp(suffix='.xlsx')
    os.close(fd)

    if _HAS_MSOFFCRYPTO:
        try:
            with open(filepath, 'rb') as f:
                office = msoffcrypto.OfficeFile(f)
                with open(temp_path, 'wb') as out:
                    office.load_key(password=password)
                    office.decrypt(out)
            return temp_path
        except msoffcrypto.InvalidKeyError:
            try:
                os.unlink(temp_path)
            except Exception:
                pass
            raise DecryptionError("密码错误")
        except Exception as e:
            err_msg = str(e).lower()
            if 'not encrypted' in err_msg or 'not an ole' in err_msg:
                shutil.copy2(filepath, temp_path)
                return temp_path
            try:
                os.unlink(temp_path)
            except Exception:
                pass
            raise DecryptionError(f"解密失败: {str(e)}")

    # 回退到内嵌实现（需加载到内存，仅小文件）
    decrypted_data = decrypt_file(filepath, password)
    with open(temp_path, 'wb') as f:
        f.write(decrypted_data)
    return temp_path


# ============================================================
# Excel 合并模块 (大文件内存优化)
# ============================================================

def _read_merged_cells_lightweight(filepath, sheet_name=None):
    """
    轻量级读取 xlsx 文件中的合并单元格信息。
    直接解析 zip 中的 XML，不加载全量数据，适合大文件。
    返回: (merged_ranges_list, max_header_row)
    """
    import zipfile
    import xml.etree.ElementTree as ET
    import re

    merged_ranges = []
    max_row = 0

    try:
        with zipfile.ZipFile(filepath, 'r') as zf:
            # 读取 workbook.xml 找到 sheet 名称和文件映射
            sheet_id_to_file = {}
            if sheet_name:
                try:
                    with zf.open('xl/workbook.xml') as wb_xml:
                        wb_root = ET.parse(wb_xml).getroot()
                    for sheet_elem in wb_root.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheet'):
                        s_name = sheet_elem.get('name', '')
                        s_rid = sheet_elem.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id', '')
                        if s_name and s_rid:
                            # 从 rels 文件找对应的文件路径
                            sheet_id_to_file[s_rid] = s_name
                except Exception:
                    pass

            # 找到工作表的 XML 文件
            sheet_files = [f for f in zf.namelist() if f.startswith('xl/worksheets/') and f.endswith('.xml')]
            if not sheet_files:
                return merged_ranges, 0

            if sheet_name and sheet_id_to_file:
                # 读取 rels 文件映射
                try:
                    with zf.open('xl/_rels/workbook.xml.rels') as rels_xml:
                        rels_root = ET.parse(rels_xml).getroot()
                    rid_to_file = {}
                    for rel in rels_root:
                        rel_id = rel.get('Id', '')
                        target = rel.get('Target', '')
                        if rel_id and target:
                            if not target.startswith('/'):
                                rid_to_file[rel_id] = 'xl/' + target
                            else:
                                rid_to_file[rel_id] = target.lstrip('/')

                    # 找到指定 sheet 的文件
                    target_file = None
                    for rid, s_name in sheet_id_to_file.items():
                        if s_name == sheet_name and rid in rid_to_file:
                            target_file = rid_to_file[rid]
                            break

                    if target_file and target_file in sheet_files:
                        sheet_files = [target_file]
                    elif target_file:
                        # 可能在 zip 中
                        for sf in sheet_files:
                            if target_file.endswith(sf.split('/')[-1]):
                                sheet_files = [sf]
                                break
                except Exception:
                    pass
            # 否则使用第一个 sheet

            with zf.open(sheet_files[0]) as sheet_xml:
                tree = ET.parse(sheet_xml)
                root = tree.getroot()

            # xlsx 命名空间
            ns = {'': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}

            # 查找 mergeCells 元素
            merge_cells_elem = root.find('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}mergeCells')
            if merge_cells_elem is None:
                return merged_ranges, 0

            for mc in merge_cells_elem:
                ref = mc.get('ref', '')
                if ref and ':' in ref:
                    merged_ranges.append(ref)
                    # 提取结束行号
                    parts = ref.split(':')
                    if len(parts) == 2:
                        digits = re.findall(r'\d+', parts[1])
                        if digits:
                            row_num = int(digits[-1])
                            if row_num > max_row:
                                max_row = row_num
    except zipfile.BadZipFile:
        # 不是 xlsx 格式（可能是 xls 或加密文件）
        pass
    except Exception:
        pass

    return merged_ranges, max_row if max_row > 0 else 0

def _read_header_row(filepath, header_rows, read_only=True, sheet_name=None):
    """
    读取文件表头的最后一行（列名行）。
    对于合并单元格，会检查整个合并区域的值。
    返回列名列表。
    """
    # read_only模式下无法正确读取合并单元格，需要用常规模式
    wb = openpyxl.load_workbook(filepath, read_only=False, data_only=True)
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb[wb.sheetnames[0]]
    
    # 构建合并单元格映射：cell -> merged_value
    merged_values = {}
    for merged_range in ws.merged_cells.ranges:
        # 获取合并区域左上角的值
        top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
        value = top_left_cell.value
        if value is not None:
            # 为合并区域中的所有单元格设置该值
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    merged_values[(row, col)] = value
    
    header_row = []
    
    # 读取指定行的数据
    row_num = header_rows
    for cell in ws[row_num]:
        # 先检查是否在合并单元格中
        if (cell.row, cell.column) in merged_values:
            header_row.append(merged_values[(cell.row, cell.column)])
        else:
            header_row.append(cell.value if cell.value is not None else '')
    
    wb.close()
    return header_row


def _build_unified_columns(all_headers):
    """
    根据所有文件的表头行构建统一列列表和映射。
    
    all_headers: [header_list1, header_list2, ...]
    返回: (unified_columns, column_mappings)
        unified_columns: 统一的列名列表
        column_mappings: [{src_idx: dst_idx, ...}, ...] 每个文件的列映射
    """
    if not all_headers:
        return [], []

    # 以第一个文件的列顺序为基础
    unified = list(all_headers[0])
    column_mappings = []

    # 为第一个文件构建映射（一一对应）
    first_mapping = {}
    for src_idx in range(len(all_headers[0])):
        first_mapping[src_idx] = src_idx
    column_mappings.append(first_mapping)

    # 处理后续文件
    for file_idx in range(1, len(all_headers)):
        header = all_headers[file_idx]
        mapping = {}
        
        for src_idx, col_name in enumerate(header):
            col_name = str(col_name).strip() if col_name else ''
            
            if not col_name:
                # 空列名：创建新的独立列（每个空列都独立）
                dst_idx = len(unified)
                unified.append('')
                mapping[src_idx] = dst_idx
            elif col_name in unified:
                # 列名已存在，映射到该位置（相同列名合并）
                dst_idx = unified.index(col_name)
                mapping[src_idx] = dst_idx
            else:
                # 新列名，追加到末尾
                dst_idx = len(unified)
                unified.append(col_name)
                mapping[src_idx] = dst_idx
        
        column_mappings.append(mapping)

    return unified, column_mappings


def _col_letter(col_index):
    """将列索引(0-based)转换为Excel列字母(A, B, ..., Z, AA, AB, ...)"""
    result = ''
    col = col_index  # 0-based
    while True:
        result = chr(65 + col % 26) + result
        col = col // 26 - 1
        if col < 0:
            break
    return result


def _expand_merged_ranges(merged_ranges_info, original_col_count, new_col_count):
    """
    扩展合并单元格范围以覆盖新增列。
    只扩展水平合并的单元格（跨多列的标题），保持垂直合并的单元格不变。
    
    判断规则：
    - 如果合并单元格跨越多列（start_col != end_col），且结束列 >= 原始最后一列，则扩展
    - 如果合并单元格只占1列（start_col == end_col），则不扩展（如E1:E2垂直合并）
    """
    if new_col_count <= original_col_count:
        return merged_ranges_info

    new_last_col = _col_letter(new_col_count - 1)
    expanded = []
    
    for mc_str in merged_ranges_info:
        parts = mc_str.split(':')
        if len(parts) == 2:
            start_ref, end_ref = parts
            
            # 提取起始列和结束列字母
            start_col = ''.join(c for c in start_ref if c.isalpha())
            end_col = ''.join(c for c in end_ref if c.isalpha())
            
            # 计算列索引（1-based）
            start_col_idx = 0
            for c in start_col:
                start_col_idx = start_col_idx * 26 + (ord(c) - 64)
            
            end_col_idx = 0
            for c in end_col:
                end_col_idx = end_col_idx * 26 + (ord(c) - 64)
            
            # 判断是否为水平合并（跨多列）
            is_horizontal_merge = (start_col_idx != end_col_idx)
            
            if is_horizontal_merge and end_col_idx >= original_col_count:
                # 水平合并且覆盖到最后一列，扩展到新的最后一列
                end_row = ''.join(c for c in end_ref if c.isdigit())
                expanded.append(f"{start_ref}:{new_last_col}{end_row}")
            else:
                # 垂直合并或未覆盖到最后一列，保持不变
                expanded.append(mc_str)
        else:
            expanded.append(mc_str)
    
    return expanded


def _inject_merge_cells_to_xlsx(xlsx_path, merged_ranges):
    """
    直接修改 xlsx 文件的 zip 内容注入合并单元格信息。
    避免用 openpyxl.load_workbook() 重新加载大文件（1GB+文件会占满内存）。
    原理：xlsx 本质是 zip 包，直接操作 sheet XML 字符串插入 mergeCells 元素。
    流式处理：对大条目使用临时文件，避免将整个文件加载到内存。
    """
    import zipfile

    if not merged_ranges:
        return

    # 构建 mergeCells XML 片段（带命名空间）
    mc_parts = [f'<mergeCell ref="{ref}"/>' for ref in merged_ranges]
    # 使用正确的命名空间
    mc_xml = ('<mergeCells xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
              f'count="{len(merged_ranges)}">' + ''.join(mc_parts) + '</mergeCells>')
    mc_xml_bytes = mc_xml.encode('utf-8')
    end_tag = b'</worksheet>'

    temp_output = xlsx_path + '.tmp'
    LARGE_ENTRY_THRESHOLD = 10 * 1024 * 1024  # 10MB
    CHUNK_SIZE = 256 * 1024  # 256KB
    first_sheet_modified = False

    try:
        with zipfile.ZipFile(xlsx_path, 'r') as zin:
            with zipfile.ZipFile(temp_output, 'w', zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    is_first_sheet = (
                        item.filename.startswith('xl/worksheets/')
                        and item.filename.endswith('.xml')
                        and not item.filename.startswith('xl/worksheets/_')
                        and not first_sheet_modified
                    )

                    if is_first_sheet:
                        # 工作表 XML：流式写临时文件，尾部插入合并单元格
                        fd, sheet_tmp = tempfile.mkstemp(suffix='.xml')
                        os.close(fd)
                        try:
                            with zin.open(item) as src, open(sheet_tmp, 'wb') as dst:
                                shutil.copyfileobj(src, dst, length=CHUNK_SIZE)

                            # 从尾部查找 </worksheet>，避免加载整个文件到内存
                            with open(sheet_tmp, 'rb') as f:
                                f.seek(0, 2)
                                file_size = f.tell()
                                tail_size = min(4096, file_size)
                                f.seek(file_size - tail_size)
                                tail = f.read()

                            tag_pos = tail.rfind(end_tag)
                            if tag_pos >= 0:
                                abs_pos = (file_size - tail_size) + tag_pos
                                with open(sheet_tmp, 'r+b') as f:
                                    f.seek(abs_pos)
                                    f.truncate()
                                    f.write(mc_xml_bytes)
                                    f.write(end_tag)

                            zout.write(sheet_tmp, item.filename,
                                       compress_type=item.compress_type)
                            first_sheet_modified = True
                        finally:
                            try:
                                os.unlink(sheet_tmp)
                            except OSError:
                                pass

                    elif item.file_size > LARGE_ENTRY_THRESHOLD:
                        # 大条目（如 sharedStrings.xml）：流式通过临时文件
                        fd, entry_tmp = tempfile.mkstemp()
                        os.close(fd)
                        try:
                            with zin.open(item) as src, open(entry_tmp, 'wb') as dst:
                                shutil.copyfileobj(src, dst, length=CHUNK_SIZE)
                            zout.write(entry_tmp, item.filename,
                                       compress_type=item.compress_type)
                        finally:
                            try:
                                os.unlink(entry_tmp)
                            except OSError:
                                pass

                    else:
                        # 小条目：直接读写
                        zout.writestr(item, zin.read(item.filename))

        shutil.move(temp_output, xlsx_path)
        logger.info(f"成功注入 {len(merged_ranges)} 个合并单元格")
    except Exception as e:
        logger.error(f"注入合并单元格失败: {str(e)}")
        # 如果失败，尝试清理临时文件
        try:
            if os.path.exists(temp_output):
                os.unlink(temp_output)
        except Exception:
            pass


@dataclass
class MergeConfig:
    """合并配置"""
    deduplicate: bool = False  # 是否去重
    skip_empty_rows: bool = True  # 跳过空行
    output_sheet_name: str = "合并数据"  # 输出Sheet名称
    max_workers: int = 4  # 并行解密线程数
    log_file: Optional[str] = None  # 日志文件路径


def _decrypt_single_file(file_info: Tuple[str, str, str, str], progress_callback=None, idx: int = 0, total: int = 1) -> Tuple[Optional[str], str, str]:
    """
    解密单个文件（用于并行处理）
    
    Returns:
        (temp_path或None, city, target_sheet)
    """
    filepath, city, password, target_sheet = file_info
    
    if progress_callback:
        sheet_info = f" [{target_sheet}]" if target_sheet else ""
        progress_callback(idx, total, f"正在解密{sheet_info}: {os.path.basename(filepath)}")
    
    if is_encrypted_file(filepath):
        if not password:
            logger.warning(f"跳过 {os.path.basename(filepath)}: 文件已加密但未提供密码")
            return (None, city, target_sheet)
        try:
            temp_path = decrypt_to_tempfile(filepath, password)
            return (temp_path, city, target_sheet)
        except DecryptionError as e:
            logger.error(f"解密失败 {os.path.basename(filepath)}: {str(e)}")
            return (None, city, target_sheet)
    else:
        return (filepath, city, target_sheet)


def merge_excel_files_with_decrypt(
    file_infos: List[Tuple[str, str, str, str]],
    output_path: str,
    progress_callback=None,
    header_rows: int = 0,
    config: Optional[MergeConfig] = None
) -> Dict[str, Any]:
    """
    合并多个加密 Excel 文件到一个工作表。
    保留第一个文件的表头（含合并单元格），后续文件跳过表头只追加数据行。
    支持不同文件有不同的额外列，按列名对齐合并。
    
    Args:
        file_infos: [(filepath, city, password, target_sheet), ...] 列表
        output_path: 输出文件路径
        progress_callback: 进度回调函数
        header_rows: 表头行数（0=自动检测，根据合并单元格范围推断）
        config: 合并配置
        
    Returns:
        合并统计信息字典
    """
    if config is None:
        config = MergeConfig()
    
    # 设置日志文件
    if config.log_file:
        file_handler = logging.FileHandler(config.log_file, encoding='utf-8')
        file_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
        logger.addHandler(file_handler)
    
    if not file_infos:
        raise MergeError("没有要合并的文件")

    total = len(file_infos)
    temp_files = []  # 解密后的临时文件
    stats = {
        'total_files': total,
        'success_files': 0,
        'failed_files': 0,
        'total_rows': 0,
        'duplicate_rows': 0,
        'total_columns': 0,
        'header_rows': 0
    }

    try:
        # 第一步：并行解密所有文件
        logger.info(f"开始解密 {total} 个文件...")
        decrypted_files = []
        
        if total > 1 and config.max_workers > 1:
            # 多文件时使用并行解密
            with ThreadPoolExecutor(max_workers=config.max_workers) as executor:
                futures = {
                    executor.submit(_decrypt_single_file, fi, None, idx, total): idx
                    for idx, fi in enumerate(file_infos)
                }
                
                for future in as_completed(futures):
                    idx = futures[future]
                    try:
                        result = future.result()
                        decrypted_files.append((idx, result))
                        if result[0]:  # temp_path
                            temp_files.append(result[0])
                            stats['success_files'] += 1
                        else:
                            stats['failed_files'] += 1
                    except Exception as e:
                        logger.error(f"解密任务异常: {str(e)}")
                        decrypted_files.append((idx, (None, file_infos[idx][1], file_infos[idx][3])))
                        stats['failed_files'] += 1
        else:
            # 单文件或小文件串行处理
            for idx, fi in enumerate(file_infos):
                result = _decrypt_single_file(fi, progress_callback, idx, total)
                decrypted_files.append((idx, result))
                if result[0]:
                    temp_files.append(result[0])
                    stats['success_files'] += 1
                else:
                    stats['failed_files'] += 1
        
        # 按原始顺序排序
        decrypted_files.sort(key=lambda x: x[0])
        decrypted_files = [df[1] for df in decrypted_files]
        
        logger.info(f"解密完成: 成功 {stats['success_files']}, 失败 {stats['failed_files']}")

        # 第二步：预扫描 - 读取合并单元格信息 + 列名信息
        detected_header_rows = header_rows
        merged_ranges_info = []

        # 2a. 读取第一个文件的合并单元格信息
        for idx_pre, (filepath_pre, city_pre, target_sheet_pre) in enumerate(decrypted_files):
            if filepath_pre is None:
                continue
            try:
                if target_sheet_pre:
                    merged_ranges_info, auto_header_rows = _read_merged_cells_lightweight(
                        filepath_pre, sheet_name=target_sheet_pre
                    )
                else:
                    merged_ranges_info, auto_header_rows = _read_merged_cells_lightweight(filepath_pre)
                if header_rows <= 0:
                    if merged_ranges_info:
                        detected_header_rows = auto_header_rows
                    if detected_header_rows <= 0:
                        detected_header_rows = 1
                if progress_callback:
                    progress_callback(0, total * 2,
                                      f"检测到表头 {detected_header_rows} 行, "
                                      f"合并单元格 {len(merged_ranges_info)} 个")
                break
            except Exception as e:
                if progress_callback:
                    progress_callback(0, total * 2,
                                      f"读取合并单元格信息失败: {str(e)}")
                detected_header_rows = max(detected_header_rows, 1)
                continue

        if detected_header_rows <= 0:
            detected_header_rows = 1

        # 2b. 读取所有文件的列名行，构建统一列映射
        all_headers = []
        valid_file_indices = []  # 记录有效文件的索引
        for idx, (filepath, city, target_sheet) in enumerate(decrypted_files):
            if filepath is None:
                all_headers.append([])
                continue
            try:
                header = _read_header_row(filepath, detected_header_rows, sheet_name=target_sheet)
                all_headers.append(header)
                valid_file_indices.append(idx)
            except Exception:
                all_headers.append([])

        unified_columns, column_mappings = _build_unified_columns(all_headers)
        total_cols = len(unified_columns)
        first_file_col_count = len(all_headers[0]) if all_headers and all_headers[0] else total_cols
        
        # 打印调试信息
        logger.info(f"统一列: {unified_columns}")
        logger.info(f"总列数: {total_cols}, 第一个文件列数: {first_file_col_count}")
        for idx, mapping in enumerate(column_mappings):
            logger.info(f"文件{idx} 列映射: {mapping}")

        # 2c. 扩展合并单元格范围（如果新增了列）
        if total_cols > first_file_col_count and merged_ranges_info:
            merged_ranges_info = _expand_merged_ranges(
                merged_ranges_info, first_file_col_count, total_cols
            )

        if progress_callback:
            extra_cols = total_cols - first_file_col_count
            progress_callback(0, total * 2,
                              f"统一列数: {total_cols}, 额外列: {extra_cols}, "
                              f"表头 {detected_header_rows} 行")

        # 第三步：合并数据到一个工作表
        output_wb = openpyxl.Workbook(write_only=True)
        out_ws = output_wb.create_sheet(title=config.output_sheet_name)

        first_file = True
        total_data_rows = 0
        seen_rows = set()  # 用于去重
        skipped_empty = 0

        for idx, (filepath, city, target_sheet) in enumerate(decrypted_files):
            if filepath is None:
                continue

            if progress_callback:
                progress_callback(total + idx, total * 2,
                                  f"正在合并: {os.path.basename(filepath)}")

            try:
                wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
                mapping = column_mappings[idx] if idx < len(column_mappings) else {}

                # 确定要读取的 sheet
                if target_sheet and target_sheet in wb.sheetnames:
                    sheets_to_read = [target_sheet]
                else:
                    sheets_to_read = wb.sheetnames

                for sheet_name in sheets_to_read:
                    ws = wb[sheet_name]

                    row_count = 0
                    for row in ws.iter_rows(values_only=True):
                        row_count += 1

                        if not first_file:
                            # 后续文件：跳过表头行
                            if row_count <= detected_header_rows:
                                continue

                        # 按列映射对齐写入
                        aligned_row = [''] * total_cols
                        
                        # 第一个文件的表头行（合并单元格行）特殊处理
                        if first_file and row_count < detected_header_rows:
                            # 合并单元格的表头行，直接按位置写入，不映射
                            cleaned_row = [cell if cell is not None else '' for cell in row]
                            # 补齐到统一列数
                            while len(cleaned_row) < total_cols:
                                cleaned_row.append('')
                            aligned_row = cleaned_row
                        else:
                            # 数据行：使用列映射对齐
                            for src_idx, cell_val in enumerate(row):
                                if src_idx in mapping:
                                    dst_idx = mapping[src_idx]
                                    if dst_idx < total_cols:
                                        # 确保不包含 None 值，write_only 模式不支持 None
                                        aligned_row[dst_idx] = cell_val if cell_val is not None else ''
                        
                        # 最后一道保险：确保所有单元格都不是 None
                        aligned_row = [cell if cell is not None else '' for cell in aligned_row]

                        # 跳过空行
                        if config.skip_empty_rows and not any(cell for cell in aligned_row):
                            skipped_empty += 1
                            continue

                        # 数据去重
                        if config.deduplicate:
                            row_tuple = tuple(aligned_row)
                            if row_tuple in seen_rows:
                                stats['duplicate_rows'] += 1
                                continue
                            seen_rows.add(row_tuple)

                        out_ws.append(aligned_row)
                        total_data_rows += 1

                        if progress_callback and total_data_rows % 2000 == 0:
                            progress_callback(total + idx, total * 2,
                                              f"合并中: {os.path.basename(filepath)} - "
                                              f"{total_data_rows} 行")

                    first_file = False

                wb.close()

            except Exception as e:
                logger.error(f"处理文件出错 {os.path.basename(filepath)}: {str(e)}")
                if progress_callback:
                    progress_callback(total + idx, total * 2,
                                      f"处理文件出错 {os.path.basename(filepath)}: {str(e)}")
                continue

        # 第四步：保存并应用合并单元格
        # 如果有合并单元格，需要用常规模式重新打开并应用
        if merged_ranges_info:
            try:
                # 先保存write_only的文件
                output_wb.save(output_path)
                
                # 然后用常规模式打开，应用合并单元格
                from openpyxl import load_workbook
                wb = load_workbook(output_path)
                ws = wb[config.output_sheet_name]
                
                # 应用合并单元格
                for mc_range in merged_ranges_info:
                    ws.merge_cells(mc_range)
                
                # 保存
                wb.save(output_path)
                wb.close()
                
                logger.info(f"成功应用 {len(merged_ranges_info)} 个合并单元格")
            except Exception as e:
                logger.error(f"应用合并单元格失败: {str(e)}")
                if progress_callback:
                    progress_callback(total * 2, total * 2,
                                      f"应用合并单元格失败: {str(e)}")
        else:
            # 没有合并单元格，直接保存
            output_wb.save(output_path)

        # 更新统计信息
        stats['total_rows'] = total_data_rows
        stats['total_columns'] = total_cols
        stats['header_rows'] = detected_header_rows
        stats['skipped_empty_rows'] = skipped_empty
        
        logger.info(
            f"合并完成! 共 {total_data_rows} 行, "
            f"{total_cols} 列, "
            f"表头 {detected_header_rows} 行, "
            f"合并单元格 {len(merged_ranges_info)} 个"
        )
        
        if progress_callback:
            progress_callback(total * 2, total * 2,
                              f"合并完成! 共 {total_data_rows} 行, "
                              f"{total_cols} 列, "
                              f"表头 {detected_header_rows} 行, "
                              f"合并单元格 {len(merged_ranges_info)} 个")

    finally:
        # 清理临时文件
        for temp_path in temp_files:
            try:
                if os.path.exists(temp_path):
                    os.unlink(temp_path)
            except Exception:
                pass
    
    return stats


# ============================================================
# GUI 界面
# ============================================================

class DropArea(QLabel):
    """文件拖拽区域"""
    files_dropped = pyqtSignal(list)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAcceptDrops(True)
        self.setMinimumHeight(50)
        self.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.setText("拖拽 Excel 文件到此处，或点击选择文件")
        self.setFont(QFont(_get_system_font(), 10))
        self.setCursor(Qt.CursorShape.PointingHandCursor)
        self._normal = "border: 2px dashed #bbb; border-radius: 6px; background: #fafafa; color: #999;"
        self._hover = "border: 2px dashed #4CAF50; border-radius: 6px; background: #e8f5e9; color: #2E7D32;"
        self.setStyleSheet(self._normal)

    def mousePressEvent(self, event):
        files, _ = QFileDialog.getOpenFileNames(
            self, "选择 Excel 文件", "",
            "Excel文件 (*.xlsx *.xls);;所有文件 (*)"
        )
        if files:
            self.files_dropped.emit(files)

    def dragEnterEvent(self, event: QDragEnterEvent):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
            self.setStyleSheet(self._hover)
        else:
            event.ignore()

    def dragLeaveEvent(self, event):
        self.setStyleSheet(self._normal)

    def dropEvent(self, event: QDropEvent):
        self.setStyleSheet(self._normal)
        files = []
        for url in event.mimeData().urls():
            filepath = url.toLocalFile()
            if filepath.lower().endswith(('.xlsx', '.xls')):
                files.append(filepath)
        if files:
            self.files_dropped.emit(files)


class MergeWorker(QThread):
    """合并工作线程"""
    progress = pyqtSignal(int, int, str)
    finished = pyqtSignal(bool, str)

    def __init__(self, file_infos, output_path, header_rows=0, config=None):
        super().__init__()
        self.file_infos = file_infos
        self.output_path = output_path
        self.header_rows = header_rows
        self.config = config if config else MergeConfig()

    def run(self):
        try:
            stats = merge_excel_files_with_decrypt(
                self.file_infos,
                self.output_path,
                progress_callback=self._progress,
                header_rows=self.header_rows,
                config=self.config
            )
            
            # 构建详细完成消息
            msg = f"合并完成!\n"
            msg += f"文件: {stats['success_files']} 成功, {stats['failed_files']} 失败\n"
            msg += f"数据: {stats['total_rows']} 行, {stats['total_columns']} 列\n"
            if stats.get('duplicate_rows', 0) > 0:
                msg += f"去重: {stats['duplicate_rows']} 行\n"
            if stats.get('skipped_empty_rows', 0) > 0:
                msg += f"跳过空行: {stats['skipped_empty_rows']} 行"
            
            self.finished.emit(True, msg)
        except Exception as e:
            self.finished.emit(False, f"合并失败: {str(e)}")

    def _progress(self, current, total, message):
        self.progress.emit(current, total, message)


class MainWindow(QMainWindow):
    """主窗口 - 简洁版"""

    # 常见城市列表（用于自动识别文件名中的城市）
    _COMMON_CITIES = [
        "北京", "上海", "广州", "深圳", "杭州", "成都",
        "武汉", "南京", "重庆", "西安", "苏州", "天津",
        "长沙", "郑州", "东莞", "青岛", "沈阳", "宁波",
        "昆明", "大连", "厦门", "福州", "珠海", "合肥",
        "济南", "佛山", "无锡", "温州", "哈尔滨",
        "石家庄", "乌鲁木齐", "呼和浩特",
        "广东", "浙江", "江苏", "四川", "湖北", "湖南",
        "福建", "山东", "河南", "河北", "辽宁", "吉林",
    ]
    
    # 配置文件路径
    _CONFIG_FILE = os.path.join(
        os.path.expanduser('~'),
        '.excel_merge_tool_config.json'
    )

    def __init__(self):
        super().__init__()
        self.setWindowTitle("加密Excel合并工具 v2.0")
        self.setMinimumSize(600, 420)
        self.resize(700, 500)

        # 数据存储
        self.file_list = []  # [(filepath, city, password, sheet_name), ...]
        self.city_passwords = OrderedDict()  # {city: password}
        self.worker = None
        self.recent_files = []  # 最近使用的文件列表
        
        # 加载配置
        self._load_config()

        self._init_ui()
        self._apply_styles()

    def _init_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        layout = QVBoxLayout(central)
        layout.setSpacing(4)
        layout.setContentsMargins(10, 10, 10, 10)

        # === 拖拽区域 ===
        self.drop_area = DropArea()
        self.drop_area.files_dropped.connect(self._on_files_dropped)
        layout.addWidget(self.drop_area)

        # === 城市密码（紧凑行内添加） ===
        pwd_row = QHBoxLayout()
        pwd_row.setSpacing(4)
        pwd_row.addWidget(QLabel("城市:"))
        self.city_input = QLineEdit()
        self.city_input.setPlaceholderText("城市名")
        self.city_input.setFixedWidth(80)
        pwd_row.addWidget(self.city_input)
        pwd_row.addWidget(QLabel("密码:"))
        self.pwd_input = QLineEdit()
        self.pwd_input.setPlaceholderText("密码")
        self.pwd_input.setEchoMode(QLineEdit.EchoMode.Password)
        self.pwd_input.setFixedWidth(120)
        pwd_row.addWidget(self.pwd_input)
        add_pwd_btn = QPushButton("添加")
        add_pwd_btn.setFixedWidth(50)
        add_pwd_btn.clicked.connect(self._add_password_inline)
        pwd_row.addWidget(add_pwd_btn)
        self.pwd_input.returnPressed.connect(self._add_password_inline)
        
        # 批量导入按钮
        import_btn = QPushButton("批量导入")
        import_btn.setFixedWidth(70)
        import_btn.clicked.connect(self._import_passwords_batch)
        pwd_row.addWidget(import_btn)
        
        # 导出按钮
        export_btn = QPushButton("导出")
        export_btn.setFixedWidth(50)
        export_btn.clicked.connect(self._export_passwords)
        pwd_row.addWidget(export_btn)
        
        pwd_row.addStretch()
        layout.addLayout(pwd_row)

        # 城市密码标签行
        self.pwd_tags_layout = QHBoxLayout()
        self.pwd_tags_layout.setSpacing(4)
        self.pwd_tags_widget = QWidget()
        self.pwd_tags_widget.setLayout(self.pwd_tags_layout)
        self.pwd_tags_widget.setFixedHeight(28)
        layout.addWidget(self.pwd_tags_widget)

        # === 文件列表 ===
        self.file_table = QTableWidget(0, 5)
        self.file_table.setHorizontalHeaderLabels(["文件名", "城市", "密码", "Sheet", "大小"])
        self.file_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.file_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Interactive)
        self.file_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Interactive)
        self.file_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Interactive)
        self.file_table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.Interactive)
        self.file_table.setColumnWidth(1, 70)
        self.file_table.setColumnWidth(2, 100)
        self.file_table.setColumnWidth(3, 90)
        self.file_table.setColumnWidth(4, 60)
        self.file_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.file_table.setEditTriggers(QAbstractItemView.EditTrigger.DoubleClicked)
        self.file_table.setItemDelegateForColumn(1, CityComboDelegate(self.file_table, self))
        self.file_table.itemChanged.connect(self._on_file_table_item_changed)
        layout.addWidget(self.file_table, 1)

        # 文件操作行
        file_btn_row = QHBoxLayout()
        file_btn_row.setSpacing(4)
        remove_btn = QPushButton("移除选中")
        remove_btn.clicked.connect(self._remove_selected_file)
        file_btn_row.addWidget(remove_btn)
        clear_btn = QPushButton("清空")
        clear_btn.clicked.connect(self._clear_file_list)
        file_btn_row.addWidget(clear_btn)
        file_btn_row.addStretch()
        layout.addLayout(file_btn_row)

        # === 底部：输出 + 合并 ===
        bottom_row = QHBoxLayout()
        bottom_row.setSpacing(4)
        bottom_row.addWidget(QLabel("输出:"))
        self.output_edit = QLineEdit()
        self.output_edit.setPlaceholderText("合并后输出文件...")
        self.output_btn = QPushButton("...")
        self.output_btn.setFixedWidth(30)
        self.output_btn.clicked.connect(self._select_output)
        bottom_row.addWidget(self.output_edit, 1)
        bottom_row.addWidget(self.output_btn)
        layout.addLayout(bottom_row)
        
        # 配置行
        config_row = QHBoxLayout()
        config_row.setSpacing(8)
        
        self.dedup_checkbox = QCheckBox("数据去重")
        self.dedup_checkbox.setToolTip("合并时去除重复数据行")
        config_row.addWidget(self.dedup_checkbox)
        
        self.skip_empty_checkbox = QCheckBox("跳过空行")
        self.skip_empty_checkbox.setChecked(True)
        self.skip_empty_checkbox.setToolTip("合并时跳过空行")
        config_row.addWidget(self.skip_empty_checkbox)
        
        config_row.addWidget(QLabel("Sheet名:"))
        self.sheet_name_input = QLineEdit("合并数据")
        self.sheet_name_input.setFixedWidth(100)
        config_row.addWidget(self.sheet_name_input)
        
        config_row.addStretch()
        layout.addLayout(config_row)

        # 进度条
        self.progress_bar = QProgressBar()
        self.progress_bar.setValue(0)
        self.progress_bar.setFixedHeight(16)
        layout.addWidget(self.progress_bar)

        # 合并按钮
        self.merge_btn = QPushButton("开始合并")
        self.merge_btn.setMinimumHeight(32)
        self.merge_btn.clicked.connect(self._start_merge)
        layout.addWidget(self.merge_btn)

        self.statusBar().showMessage("就绪")

    def _apply_styles(self):
        self.setStyleSheet("""
            QMainWindow { background: #fff; }
            QPushButton {
                background: #2196F3; color: white; border: none;
                border-radius: 3px; padding: 4px 10px; font-size: 11px;
            }
            QPushButton:hover { background: #1976D2; }
            QPushButton:disabled { background: #bbb; }
            QPushButton#mergeBtn {
                background-color: #4CAF50; font-weight: bold;
                font-size: 13px; border-radius: 4px;
            }
            QPushButton#mergeBtn:hover { background-color: #388E3C; }
            QPushButton#mergeBtn:disabled { background-color: #ccc; }
            QTableWidget { border: 1px solid #ddd; gridline-color: #eee; }
            QHeaderView::section {
                background: #f5f5f5; border: none;
                border-bottom: 1px solid #ddd; padding: 3px; font-weight: bold;
            }
            QProgressBar {
                border: 1px solid #ddd; border-radius: 3px;
                text-align: center; height: 14px;
            }
            QProgressBar::chunk { background: #4CAF50; border-radius: 2px; }
        """)
        self.merge_btn.setObjectName("mergeBtn")

    # ---- 城市密码管理 ----

    def _add_password_inline(self):
        """从输入框添加城市密码"""
        city = self.city_input.text().strip()
        password = self.pwd_input.text()
        if not city:
            return
        self.city_passwords[city] = password
        self.city_input.clear()
        self.pwd_input.clear()
        self._refresh_pwd_tags()
        self._auto_match_all_files()
        self.statusBar().showMessage(f"已添加: {city}")

    def _remove_city_password(self, city):
        """删除城市密码"""
        if city in self.city_passwords:
            del self.city_passwords[city]
            self._refresh_pwd_tags()
            self._auto_match_all_files()

    def _refresh_pwd_tags(self):
        """刷新城市密码标签"""
        # 清除旧标签
        while self.pwd_tags_layout.count():
            item = self.pwd_tags_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        if not self.city_passwords:
            lbl = QLabel("(未配置城市密码)")
            lbl.setStyleSheet("color: #aaa; font-size: 10px;")
            self.pwd_tags_layout.addWidget(lbl)
            self.pwd_tags_layout.addStretch()
            return

        for city, password in self.city_passwords.items():
            tag_btn = QPushButton(f"{city}: {'*' * min(len(password), 6)}")
            tag_btn.setStyleSheet(
                "background: #e3f2fd; color: #1565C0; border: 1px solid #90CAF9; "
                "border-radius: 3px; padding: 2px 6px; font-size: 10px;"
            )
            tag_btn.setFixedHeight(22)
            tag_btn.setToolTip(f"点击删除 {city}")
            tag_btn.clicked.connect(lambda checked, c=city: self._remove_city_password(c))
            self.pwd_tags_layout.addWidget(tag_btn)
        self.pwd_tags_layout.addStretch()

    def _auto_match_all_files(self):
        """重新扫描所有文件，根据城市密码自动匹配密码"""
        self.file_table.itemChanged.disconnect(self._on_file_table_item_changed)
        try:
            for row in range(self.file_table.rowCount()):
                city_item = self.file_table.item(row, 1)
                pwd_item = self.file_table.item(row, 2)
                if not city_item or not pwd_item:
                    continue
                
                city = city_item.text().strip()
                
                # 如果城市为空，尝试从文件名重新识别
                if not city and row < len(self.file_list):
                    filepath = self.file_list[row][0]
                    city = self._guess_city_from_filename(filepath)
                    city_item.setText(city)
                
                # 根据城市自动填充密码
                if city and city in self.city_passwords:
                    new_pwd = self.city_passwords[city]
                    if pwd_item.text() != new_pwd:
                        pwd_item.setText(new_pwd)
                        if row < len(self.file_list):
                            self.file_list[row] = (self.file_list[row][0], city, new_pwd, self.file_list[row][3])
                elif city:
                    # 城市不在密码列表中，清空密码
                    if pwd_item.text():
                        pwd_item.setText("")
                        if row < len(self.file_list):
                            self.file_list[row] = (self.file_list[row][0], city, "", self.file_list[row][3])
        finally:
            self.file_table.itemChanged.connect(self._on_file_table_item_changed)

    # ---- 文件管理 ----

    def _on_files_dropped(self, files):
        """处理拖拽/选择的文件"""
        added = 0
        for filepath in files:
            existing = [f[0] for f in self.file_list]
            if filepath in existing:
                continue

            size_str = self._format_size(os.path.getsize(filepath))
            city = self._guess_city_from_filename(filepath)
            password = self.city_passwords.get(city, "")

            self.file_list.append((filepath, city, password, ""))
            row = self.file_table.rowCount()
            self.file_table.insertRow(row)
            self.file_table.setItem(row, 0, QTableWidgetItem(os.path.basename(filepath)))
            self.file_table.setItem(row, 1, QTableWidgetItem(city))
            self.file_table.setItem(row, 2, QTableWidgetItem(password))
            self.file_table.setItem(row, 3, QTableWidgetItem(""))  # sheet
            self.file_table.setItem(row, 4, QTableWidgetItem(size_str))

            if is_encrypted_file(filepath):
                self.file_table.item(row, 0).setBackground(QColor(255, 243, 224))

            added += 1

        self.statusBar().showMessage(f"已添加 {added} 个文件，共 {len(self.file_list)} 个")

    def _guess_city_from_filename(self, filepath):
        """从文件名推断城市"""
        name = os.path.splitext(os.path.basename(filepath))[0]
        for city in sorted(self.city_passwords.keys(), key=len, reverse=True):
            if city in name:
                return city
        for city in sorted(self._COMMON_CITIES, key=len, reverse=True):
            if city in name:
                if city not in self.city_passwords:
                    self.city_passwords[city] = ""
                    self._refresh_pwd_tags()
                return city
        return ""

    def _format_size(self, size):
        """格式化文件大小"""
        for unit in ['B', 'KB', 'MB', 'GB']:
            if size < 1024:
                return f"{size:.1f}{unit}"
            size /= 1024
        return f"{size:.1f}TB"

    def _on_file_table_item_changed(self, item):
        """文件表格内容变化时，如果城市列变化则自动填充对应密码"""
        if item is None or item.column() != 1:
            return
        row = item.row()
        city = item.text().strip()
        pwd_item = self.file_table.item(row, 2)
        if pwd_item and city and city in self.city_passwords:
            pwd = self.city_passwords[city]
            pwd_item.setText(pwd)
            if row < len(self.file_list):
                sheet = self.file_list[row][3] if len(self.file_list[row]) > 3 else ""
                self.file_list[row] = (self.file_list[row][0], city, pwd, sheet)
        elif pwd_item:
            pwd_item.setText("")
            if row < len(self.file_list):
                sheet = self.file_list[row][3] if len(self.file_list[row]) > 3 else ""
                self.file_list[row] = (self.file_list[row][0], city, "", sheet)

    def _clear_file_list(self):
        """清空文件列表"""
        self.file_list.clear()
        self.file_table.setRowCount(0)
        self.statusBar().showMessage("已清空")

    def _remove_selected_file(self):
        """移除选中的文件"""
        rows = set()
        for item in self.file_table.selectedItems():
            rows.add(item.row())
        for row in sorted(rows, reverse=True):
            if row < len(self.file_list):
                self.file_list.pop(row)
            self.file_table.removeRow(row)

    # ---- 合并 ----

    def _select_output(self):
        """选择输出文件"""
        filepath, _ = QFileDialog.getSaveFileName(
            self, "选择输出文件", "merged_output.xlsx", "Excel文件 (*.xlsx)"
        )
        if filepath:
            self.output_edit.setText(filepath)

    def _collect_file_infos(self):
        """收集所有文件信息"""
        for row in range(self.file_table.rowCount()):
            if row < len(self.file_list):
                city_item = self.file_table.item(row, 1)
                pwd_item = self.file_table.item(row, 2)
                sheet_item = self.file_table.item(row, 3)
                city = city_item.text() if city_item else ""
                password = pwd_item.text() if pwd_item else ""
                sheet_name = sheet_item.text().strip() if sheet_item else ""
                self.file_list[row] = (self.file_list[row][0], city, password, sheet_name)

        file_infos = []
        for filepath, city, password, sheet_name in self.file_list:
            file_infos.append((filepath, city, password, sheet_name))
        return file_infos

    def _start_merge(self):
        """开始合并"""
        if not self.file_list:
            QMessageBox.warning(self, "提示", "请先添加要合并的 Excel 文件")
            return

        output_path = self.output_edit.text().strip()
        if not output_path:
            output_path, _ = QFileDialog.getSaveFileName(
                self, "选择输出文件", "merged_output.xlsx", "Excel文件 (*.xlsx)"
            )
            if not output_path:
                return
            self.output_edit.setText(output_path)

        file_infos = self._collect_file_infos()
        
        # 构建合并配置
        config = MergeConfig(
            deduplicate=self.dedup_checkbox.isChecked(),
            skip_empty_rows=self.skip_empty_checkbox.isChecked(),
            output_sheet_name=self.sheet_name_input.text().strip() or "合并数据",
            max_workers=4,
            log_file=None  # 可以添加日志文件选择
        )

        # 检查加密文件缺密码
        missing_pwd = [os.path.basename(fp) for fp, city, pwd, _ in file_infos
                       if is_encrypted_file(fp) and not pwd]
        if missing_pwd:
            reply = QMessageBox.question(
                self, "密码缺失",
                f"以下加密文件未设置密码:\n{chr(10).join(missing_pwd)}\n\n"
                f"未设置密码的文件将被跳过，是否继续?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            if reply == QMessageBox.StandardButton.No:
                return

        # 检查大文件
        total_size = sum(os.path.getsize(fp) for fp, _, _, _ in file_infos if os.path.exists(fp))
        if total_size > 500 * 1024 * 1024 and not _HAS_MSOFFCRYPTO:
            encrypted_any = any(is_encrypted_file(fp) for fp, _, _, _ in file_infos)
            if encrypted_any:
                QMessageBox.warning(
                    self, "性能警告",
                    f"文件总大小 {self._format_size(total_size)}，\n"
                    f"未安装 msoffcrypto-tool 库，大文件解密可能占用大量内存。\n"
                    f"建议运行: pip install msoffcrypto-tool"
                )

        self.merge_btn.setEnabled(False)
        self.merge_btn.setText("合并中...")
        self.progress_bar.setValue(0)

        self.worker = MergeWorker(file_infos, output_path, header_rows=0, config=config)
        self.worker.progress.connect(self._on_progress)
        self.worker.finished.connect(self._on_merge_finished)
        self.worker.start()

    def _on_progress(self, current, total, message):
        """进度更新"""
        if total > 0:
            self.progress_bar.setValue(int(current / total * 100))
        self.statusBar().showMessage(message)

    def _on_merge_finished(self, success, message):
        """合并完成"""
        self.merge_btn.setEnabled(True)
        self.merge_btn.setText("开始合并")

        if success:
            self.progress_bar.setValue(100)
            self.statusBar().showMessage("合并完成!")
            QMessageBox.information(self, "完成", f"合并成功!\n{self.output_edit.text()}")
        else:
            self.statusBar().showMessage("合并失败")
            QMessageBox.critical(self, "失败", message)

    def closeEvent(self, event):
        """关闭窗口"""
        if self.worker and self.worker.isRunning():
            reply = QMessageBox.question(
                self, "确认退出", "正在合并中，确定退出?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            if reply == QMessageBox.StandardButton.No:
                event.ignore()
                return
            self.worker.terminate()
            self.worker.wait()
        
        # 保存配置
        self._save_config()
        event.accept()
    
    def _load_config(self):
        """加载用户配置"""
        try:
            if os.path.exists(self._CONFIG_FILE):
                with open(self._CONFIG_FILE, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                
                # 加载城市密码
                if 'city_passwords' in config:
                    self.city_passwords = OrderedDict(config['city_passwords'])
                
                # 加载最近文件
                if 'recent_files' in config:
                    self.recent_files = config['recent_files'][:10]  # 最多保留10个
                    
                logger.info(f"配置加载成功: {len(self.city_passwords)} 个城市密码")
        except Exception as e:
            logger.warning(f"加载配置失败: {str(e)}")
    
    def _save_config(self):
        """保存用户配置"""
        try:
            config = {
                'city_passwords': dict(self.city_passwords),
                'recent_files': self.recent_files
            }
            
            with open(self._CONFIG_FILE, 'w', encoding='utf-8') as f:
                json.dump(config, f, ensure_ascii=False, indent=2)
            
            logger.info(f"配置已保存: {len(self.city_passwords)} 个城市密码")
        except Exception as e:
            logger.error(f"保存配置失败: {str(e)}")
    
    def _import_passwords_batch(self):
        """批量导入城市密码（从 Excel、CSV 或 JSON 文件）"""
        filepath, _ = QFileDialog.getOpenFileName(
            self, "导入城市密码", "",
            "Excel/CSV/JSON文件 (*.xlsx *.xls *.csv *.json);;所有文件 (*)"
        )
        
        if not filepath:
            return
        
        try:
            imported_count = 0
            
            if filepath.endswith('.json'):
                # JSON 格式: [{"city": "北京", "password": "123"}, ...]
                with open(filepath, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                
                if isinstance(data, list):
                    for item in data:
                        if isinstance(item, dict) and 'city' in item:
                            city = item['city']
                            password = item.get('password', '')
                            self.city_passwords[city] = password
                            imported_count += 1
                elif isinstance(data, dict):
                    # 直接是 {"city": "password"} 格式
                    for city, password in data.items():
                        self.city_passwords[city] = password
                        imported_count += 1
            
            elif filepath.endswith('.csv'):
                # CSV 格式: city,password
                with open(filepath, 'r', encoding='utf-8') as f:
                    reader = csv.reader(f)
                    header = next(reader, None)  # 跳过表头
                    
                    for row in reader:
                        if len(row) >= 2:
                            city = row[0].strip()
                            password = row[1].strip()
                            if city:
                                self.city_passwords[city] = password
                                imported_count += 1
            
            elif filepath.endswith(('.xlsx', '.xls')):
                # Excel 格式: 第一列城市名称，第二列密码
                wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
                ws = wb.active
                
                # 跳过表头（如果有的话），从第二行开始读取
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if len(row) >= 2:
                        city = str(row[0]).strip() if row[0] is not None else ''
                        password = str(row[1]).strip() if row[1] is not None else ''
                        if city:
                            self.city_passwords[city] = password
                            imported_count += 1
                
                wb.close()
            
            self._refresh_pwd_tags()
            self._auto_match_all_files()
            self._save_config()
            
            QMessageBox.information(
                self, "导入成功",
                f"成功导入 {imported_count} 个城市密码"
            )
            
        except Exception as e:
            QMessageBox.critical(self, "导入失败", f"导入失败: {str(e)}")
    
    def _export_passwords(self):
        """导出城市密码到 JSON 或 CSV"""
        if not self.city_passwords:
            QMessageBox.warning(self, "提示", "没有可导出的城市密码")
            return
        
        filepath, _ = QFileDialog.getSaveFileName(
            self, "导出城市密码", "city_passwords.json",
            "JSON文件 (*.json);;CSV文件 (*.csv);;所有文件 (*)"
        )
        
        if not filepath:
            return
        
        try:
            if filepath.endswith('.json'):
                # JSON 格式
                data = [
                    {"city": city, "password": pwd}
                    for city, pwd in self.city_passwords.items()
                ]
                with open(filepath, 'w', encoding='utf-8') as f:
                    json.dump(data, f, ensure_ascii=False, indent=2)
            
            elif filepath.endswith('.csv'):
                # CSV 格式
                with open(filepath, 'w', encoding='utf-8', newline='') as f:
                    writer = csv.writer(f)
                    writer.writerow(['city', 'password'])
                    for city, pwd in self.city_passwords.items():
                        writer.writerow([city, pwd])
            
            QMessageBox.information(
                self, "导出成功",
                f"成功导出 {len(self.city_passwords)} 个城市密码到:\n{filepath}"
            )
            
        except Exception as e:
            QMessageBox.critical(self, "导出失败", f"导出失败: {str(e)}")


class CityComboDelegate(QStyledItemDelegate):
    """城市列的下拉框代理"""

    def __init__(self, parent, main_window):
        super().__init__(parent)
        self.main_window = main_window

    def createEditor(self, parent, option, index):
        combo = QComboBox(parent)
        combo.setEditable(True)
        combo.addItems(list(self.main_window.city_passwords.keys()))
        # 添加一些常见城市
        common_cities = [
            "北京", "上海", "广州", "深圳", "杭州", "成都",
            "武汉", "南京", "重庆", "西安", "苏州", "天津",
            "长沙", "郑州", "东莞", "青岛", "沈阳", "宁波",
            "昆明", "大连", "厦门", "福州", "珠海", "合肥"
        ]
        for city in common_cities:
            if combo.findText(city) < 0:
                combo.addItem(city)
        return combo

    def setEditorData(self, editor, index):
        text = index.data(Qt.ItemDataRole.DisplayRole)
        idx = editor.findText(text)
        if idx >= 0:
            editor.setCurrentIndex(idx)
        else:
            editor.setEditText(text)

    def setModelData(self, editor, model, index):
        model.setData(index, editor.currentText(), Qt.ItemDataRole.EditRole)

    def updateEditorGeometry(self, editor, option, index):
        editor.setGeometry(option.rect)


# ============================================================
# 程序入口
# ============================================================

def run_cli(args):
    """
    命令行模式
    
    用法:
        python main.py --cli \
            --files file1.xlsx file2.xlsx \
            --passwords '{"北京": "123", "上海": "456"}' \
            --output merged.xlsx \
            --dedup \
            --sheet-name "合并数据" \
            --log merge.log
    """
    import argparse
    
    parser = argparse.ArgumentParser(description='加密Excel合并工具 - 命令行模式')
    parser.add_argument('--files', nargs='+', required=True, help='要合并的Excel文件列表')
    parser.add_argument('--passwords', type=str, default='{}', help='城市密码JSON字符串')
    parser.add_argument('--password-file', type=str, help='从文件加载城市密码(JSON或CSV)')
    parser.add_argument('--output', type=str, default='merged_output.xlsx', help='输出文件路径')
    parser.add_argument('--dedup', action='store_true', help='启用数据去重')
    parser.add_argument('--skip-empty', action='store_true', default=True, help='跳过空行(默认启用)')
    parser.add_argument('--no-skip-empty', action='store_true', help='不跳过空行')
    parser.add_argument('--sheet-name', type=str, default='合并数据', help='输出Sheet名称')
    parser.add_argument('--header-rows', type=int, default=0, help='表头行数(0=自动检测)')
    parser.add_argument('--log', type=str, help='日志文件路径')
    parser.add_argument('--workers', type=int, default=4, help='并行解密线程数')
    
    parsed = parser.parse_args(args)
    
    # 加载密码
    city_passwords = {}
    
    # 从 JSON 字符串加载
    try:
        city_passwords = json.loads(parsed.passwords)
    except json.JSONDecodeError as e:
        print(f"警告: 解析密码JSON失败: {str(e)}")
    
    # 从文件加载
    if parsed.password_file:
        try:
            if parsed.password_file.endswith('.json'):
                with open(parsed.password_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    if isinstance(data, dict):
                        city_passwords.update(data)
                    elif isinstance(data, list):
                        for item in data:
                            if isinstance(item, dict) and 'city' in item:
                                city_passwords[item['city']] = item.get('password', '')
            elif parsed.password_file.endswith('.csv'):
                with open(parsed.password_file, 'r', encoding='utf-8') as f:
                    reader = csv.reader(f)
                    next(reader, None)  # 跳过表头
                    for row in reader:
                        if len(row) >= 2:
                            city_passwords[row[0].strip()] = row[1].strip()
            print(f"从文件加载了 {len(city_passwords)} 个城市密码")
        except Exception as e:
            print(f"警告: 加载密码文件失败: {str(e)}")
    
    # 构建文件信息
    file_infos = []
    for filepath in parsed.files:
        if not os.path.exists(filepath):
            print(f"警告: 文件不存在: {filepath}")
            continue
        
        # 自动匹配城市密码
        city = ""
        password = ""
        filename = os.path.basename(filepath)
        
        for c, p in city_passwords.items():
            if c in filename:
                city = c
                password = p
                break
        
        file_infos.append((filepath, city, password, ""))
    
    if not file_infos:
        print("错误: 没有有效的文件")
        sys.exit(1)
    
    # 构建配置
    config = MergeConfig(
        deduplicate=parsed.dedup,
        skip_empty_rows=not parsed.no_skip_empty,
        output_sheet_name=parsed.sheet_name,
        max_workers=parsed.workers,
        log_file=parsed.log
    )
    
    # 执行合并
    print(f"\n开始合并 {len(file_infos)} 个文件...")
    print(f"输出: {parsed.output}")
    print(f"配置: 去重={config.deduplicate}, 跳过空行={config.skip_empty_rows}, Sheet={config.output_sheet_name}\n")
    
    def progress_callback(current, total, message):
        print(f"[{current}/{total}] {message}")
    
    try:
        stats = merge_excel_files_with_decrypt(
            file_infos,
            parsed.output,
            progress_callback=progress_callback,
            header_rows=parsed.header_rows,
            config=config
        )
        
        print(f"\n✓ 合并成功!")
        print(f"  文件: {stats['success_files']} 成功, {stats['failed_files']} 失败")
        print(f"  数据: {stats['total_rows']} 行, {stats['total_columns']} 列")
        print(f"  表头: {stats['header_rows']} 行")
        if stats.get('duplicate_rows', 0) > 0:
            print(f"  去重: {stats['duplicate_rows']} 行")
        if stats.get('skipped_empty_rows', 0) > 0:
            print(f"  跳过空行: {stats['skipped_empty_rows']} 行")
        print(f"  输出: {parsed.output}\n")
        
    except Exception as e:
        print(f"\n✗ 合并失败: {str(e)}\n")
        sys.exit(1)


def main():
    # 检查是否是命令行模式
    if '--cli' in sys.argv:
        # 移除 --cli 参数
        args = [arg for arg in sys.argv[1:] if arg != '--cli']
        run_cli(args)
        return
    
    # GUI 模式
    app = QApplication(sys.argv)
    app.setStyle("Fusion")

    font = QFont(_get_system_font(), 10)
    app.setFont(font)

    window = MainWindow()
    window.show()
    sys.exit(app.exec())


if __name__ == '__main__':
    main()
